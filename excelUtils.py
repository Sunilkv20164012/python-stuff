import openpyxl
from openpyxl import Workbook
import csv

def readCSV(file_name):
    print("Reading CSV content", file_name)
    with open(file_name, mode='r') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        for row in csv_reader:
            yield [element for element in row]
        print("Completed Reading CSV file ", file_name)



def writeCSV(content, header, file_name):
    print("Got CSV content", content, file_name)
    with open(file_name, mode='w') as csv_file:
        
        # writer.writeheader()
        # fieldnames = [header]
        # writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
        # # Content is Json object/Dictionary object
        # writer.writerow(content)
        csv_writer = csv.writer(csv_file, delimiter=',')
        csv_writer.writerows(content)



def convertCSVtoXLSX(inputFile, file_name):
    print("convertCSVtoXLSX CSV content", inputFile, file_name)
    with open(inputFile, mode='r') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        
        createXLSX(file_name)
        wb_obj = openpyxl.load_workbook(filename = file_name)
        sheet_obj = wb_obj.active
        for row in csv_reader:
            sheet_obj.append(row)
            print(row)
        wb_obj.save(file_name)
    
 
    
def readXLSX(file_name):

    print("Reading xlsx file  .... ", file_name)
    wb_obj = openpyxl.load_workbook(filename = file_name)
    sheet_obj = wb_obj.active

    # Loop will print(all columns name
    # for row in sheet_obj:
    #     print(row

    # for i in range(1, max_row + 1):
    #     for j in range(1, max_col + 1):
    #         cell_obj = sheet_obj.cell(row = i, column = j)
    #         print(cell_obj.value, ", ",
    #     print("\n "
    for row in sheet_obj.iter_rows():
        yield [cell.value for cell in row]
    print("Completed Reading xlsx file ", file_name)
        

        



def readColumn(colm, file_name):

    print("Reading column  .... ", colm)
    wb_obj = openpyxl.load_workbook(filename = file_name)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row
    # Loop will print(all columns name
    for i in range(1, max_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = colm)
        print("column ", i, cell_obj.value )


def createXLSX(file_name):
    print("Creating sheet .... ")
    wb_obj = openpyxl.Workbook(write_only=True)
    ws = wb_obj.create_sheet("Mysheet0")
    wb_obj.save(file_name)
    print("Sheet Created  .... ", file_name)


def writeXLSX(content, file_name):    
    createXLSX(file_name)
    appendXLSX(content, file_name)



def appendXLSX(content, file_name):
    print("appendXLSX sheet .... ")
    wb_obj = openpyxl.load_workbook(filename = file_name)
    sheet_obj = wb_obj.active
    for element in content:
        print(element)
        sheet_obj.append([element])
    wb_obj.save(file_name)
    print("Completed appendXLSX sheet ")


def getInfoXLSX(file_name):
    wb_obj = openpyxl.load_workbook(filename = file_name)
    sheet_obj = wb_obj.active
    
    print("the total number of rows", sheet_obj.max_row)
    print("total number of column", sheet_obj.max_column)
    print("sheetnames ", wb_obj.sheetnames)





